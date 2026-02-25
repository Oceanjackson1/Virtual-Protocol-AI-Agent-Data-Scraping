"""
Excel exporter: single-sheet, two-level merged headers, priority-ordered columns.
Headers in English, content in Chinese where applicable.
"""

import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import AgentData, GlobalMetrics, Offering

HEADER_STRUCTURE = [
    # (level1_name, [(level2_name, field_key), ...])
    ("Core Info", [
        ("Rank", "rank"),
        ("Agent Link", "agent_link"),
        ("Name", "name"),
        ("Category", "category"),
        ("Description", "description"),
    ]),
    ("Key Metrics", [
        ("Volume (Total AGDP)", "volume"),
        ("Gross AGDP", "gross_agdp"),
        ("Total Revenue", "revenue"),
        ("Success Rate (%)", "success_rate"),
        ("Rating", "rating"),
    ]),
    ("Activity", [
        ("Transaction Count", "transaction_count"),
        ("Successful Jobs", "successful_jobs"),
        ("Unique Buyers", "unique_buyers"),
        ("Online Status", "online_status"),
        ("Last Active", "last_active_at"),
    ]),
    ("What I Offer", [
        ("Offering Names", "offering_names"),
        ("Offering Descriptions", "offering_descs"),
        ("Offering Prices", "offering_prices"),
        ("Offering SLA (min)", "offering_sla"),
        ("Offering Requirements", "offering_reqs"),
    ]),
    ("Identity & Links", [
        ("Wallet Address", "wallet_address"),
        ("Contract Address", "contract_address"),
        ("Token Address", "token_address"),
        ("Owner Address", "owner_address"),
        ("Twitter Handle", "twitter_handle"),
        ("Symbol", "symbol"),
        ("Role", "role"),
        ("Cluster", "cluster"),
        ("Graduated", "has_graduated"),
        ("Wallet Balance", "wallet_balance"),
        ("Enabled Chains", "enabled_chains"),
        ("Virtual Agent ID", "virtual_agent_id"),
        ("Is Virtual Agent", "is_virtual_agent"),
        ("Created At", "created_at"),
        ("Profile Pic URL", "profile_pic_url"),
    ]),
]

L1_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
L1_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
L2_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
L2_FONT = Font(name="Arial", bold=True, color="1F3864", size=10)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def _format_offerings_field(offerings: List[Offering], field: str) -> str:
    if not offerings:
        return "无"
    lines = []
    for i, o in enumerate(offerings, 1):
        if field == "names":
            lines.append(f"{i}. {o.name}")
        elif field == "descs":
            lines.append(f"{i}. {o.description}" if o.description else f"{i}. 无描述")
        elif field == "prices":
            if o.price_type == "percentage":
                lines.append(f"{i}. {o.price * 100:.1f}% (按比例)")
            else:
                lines.append(f"{i}. ${o.price:.2f} USDC (固定价格)")
        elif field == "sla":
            lines.append(f"{i}. {o.sla_minutes} 分钟")
        elif field == "reqs":
            lines.append(f"{i}. {o.requirement}" if o.requirement else f"{i}. 无要求")
    return "\n".join(lines)


LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")


def _get_cell_value(agent: AgentData, key: str):
    if key == "offering_names":
        return _format_offerings_field(agent.offerings, "names")
    elif key == "offering_descs":
        return _format_offerings_field(agent.offerings, "descs")
    elif key == "offering_prices":
        return _format_offerings_field(agent.offerings, "prices")
    elif key == "offering_sla":
        return _format_offerings_field(agent.offerings, "sla")
    elif key == "offering_reqs":
        return _format_offerings_field(agent.offerings, "reqs")
    elif key == "has_graduated":
        return "是" if agent.has_graduated else "否"
    elif key == "is_virtual_agent":
        return "是" if agent.is_virtual_agent else "否"
    elif key == "success_rate":
        return agent.success_rate
    elif key == "twitter_handle":
        return f"@{agent.twitter_handle}" if agent.twitter_handle else ""
    elif key == "agent_link":
        return agent.agent_link
    else:
        return getattr(agent, key, "")


def _auto_col_widths(ws, col_start: int, col_end: int, max_width: int = 40):
    for col_idx in range(col_start, col_end + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=min(ws.max_row, 50)):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    line_max = max(len(line) for line in lines) if lines else 0
                    max_len = max(max_len, line_max)
        header_cell = ws.cell(row=2, column=col_idx)
        if header_cell.value:
            max_len = max(max_len, len(str(header_cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_width)


def export_to_excel(
    agents: List[AgentData],
    global_metrics: GlobalMetrics,
    output_dir: str = "./output",
    filename_prefix: str = "acp_agents",
) -> str:
    """Export agent data to a single-sheet Excel with two-level headers."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"{filename_prefix}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "ACP Agents"

    # -- Row 1: Level-1 headers (merged cells) --
    col = 1
    for l1_name, l2_fields in HEADER_STRUCTURE:
        span = len(l2_fields)
        start_col = col
        end_col = col + span - 1
        if span > 1:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )
        cell = ws.cell(row=1, column=start_col, value=l1_name)
        cell.font = L1_FONT
        cell.fill = L1_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        for c in range(start_col + 1, end_col + 1):
            sc = ws.cell(row=1, column=c)
            sc.fill = L1_FILL
            sc.border = THIN_BORDER
        col = end_col + 1

    # -- Row 2: Level-2 headers --
    col = 1
    for _, l2_fields in HEADER_STRUCTURE:
        for l2_name, _ in l2_fields:
            cell = ws.cell(row=2, column=col, value=l2_name)
            cell.font = L2_FONT
            cell.fill = L2_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            col += 1

    # -- Data rows (starting row 3) --
    for row_idx, agent in enumerate(agents, 3):
        col = 1
        for _, l2_fields in HEADER_STRUCTURE:
            for _, field_key in l2_fields:
                value = _get_cell_value(agent, field_key)
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                if field_key == "agent_link" and agent.agent_link:
                    cell.hyperlink = agent.agent_link
                    cell.value = agent.agent_link
                    cell.font = LINK_FONT
                else:
                    cell.font = DATA_FONT
                col += 1

    # -- Formatting --
    total_cols = sum(len(fields) for _, fields in HEADER_STRUCTURE)

    ws.freeze_panes = "A3"

    _auto_col_widths(ws, 1, total_cols)

    ws.sheet_properties.filterMode = False
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{ws.max_row}"

    # -- Summary row at the bottom --
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="爬取时间").font = Font(bold=True, size=10)
    ws.cell(row=summary_row, column=2, value=global_metrics.scrape_time).font = DATA_FONT
    ws.cell(row=summary_row + 1, column=1, value="总 Agent 数量").font = Font(bold=True, size=10)
    ws.cell(row=summary_row + 1, column=2, value=global_metrics.total_agents).font = DATA_FONT
    ws.cell(row=summary_row + 2, column=1, value="平台总 AGDP").font = Font(bold=True, size=10)
    ws.cell(row=summary_row + 2, column=2, value=f"${global_metrics.total_agdp_latest:,.2f}").font = DATA_FONT

    wb.save(filepath)
    return filepath
